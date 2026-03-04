"""
ExcelReporter Worker
リサーチ結果をExcelレポートにまとめる
"""
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import settings
from src.agents.base_agent import BaseAgent


class ExcelReporter(BaseAgent):
    """Excelレポート生成 Worker"""

    def __init__(self):
        super().__init__(
            name="ExcelReporter",
            role="リサーチ結果のExcelレポート生成担当",
            model=settings.sonnet_model,
            system_prompt="""あなたはレポート作成の専門家です。

役割:
- リサーチ結果を整理してわかりやすいレポートを作成
- データの要約と分析コメントを追加

出力形式:
- JSON形式で分析結果を返す
"""
        )

    def execute(self, context: Dict[str, Any]) -> Dict[str, Any]:
        """Excelレポート生成タスクを実行"""
        logger.info(f"[{self.name}] レポート生成開始")

        articles = context.get("articles", [])
        theme = context.get("theme", "マーケティングリサーチ")
        selected_articles = context.get("selected_articles", articles)

        if not selected_articles:
            return {
                "status": "error",
                "message": "レポート対象の記事がありません",
            }

        # Step 1: LLMで分析サマリーを生成
        analysis = self._generate_analysis(selected_articles, theme)

        # Step 2: Excelファイルを作成
        output_path = self._create_excel_report(selected_articles, analysis, theme)

        return {
            "status": "success",
            "output_path": str(output_path),
            "article_count": len(selected_articles),
            "analysis": analysis,
            "created_at": datetime.now().isoformat(),
        }

    def _generate_analysis(
        self, articles: List[Dict[str, Any]], theme: str
    ) -> Dict[str, Any]:
        """LLMでリサーチ結果を分析"""
        import json

        articles_summary = json.dumps(
            [
                {
                    "title": a.get("original_title", a.get("title", "")),
                    "category": a.get("category", ""),
                    "relevance_score": a.get("relevance_score", 0),
                    "japan_value": a.get("japan_value", ""),
                }
                for a in articles
            ],
            ensure_ascii=False,
        )

        prompt = f"""以下のリサーチ結果を分析し、サマリーを作成してください。

テーマ: {theme}
記事数: {len(articles)}

記事概要:
{articles_summary}

以下の形式で出力してください:
{{
    "summary": "全体サマリー（100文字以内）",
    "key_trends": ["トレンド1", "トレンド2", "トレンド3"],
    "recommendations": ["推奨アクション1", "推奨アクション2"],
    "category_breakdown": {{"カテゴリ名": 件数, ...}}
}}"""

        try:
            return self.call_llm_json(prompt)
        except Exception as e:
            logger.error(f"[{self.name}] 分析生成エラー: {e}")
            return {
                "summary": "分析を生成できませんでした",
                "key_trends": [],
                "recommendations": [],
                "category_breakdown": {},
            }

    def _create_excel_report(
        self,
        articles: List[Dict[str, Any]],
        analysis: Dict[str, Any],
        theme: str,
    ) -> Path:
        """Excelレポートを作成"""
        wb = Workbook()

        # サマリーシート
        ws_summary = wb.active
        ws_summary.title = "サマリー"
        self._create_summary_sheet(ws_summary, analysis, theme, len(articles))

        # 記事一覧シート
        ws_articles = wb.create_sheet("記事一覧")
        self._create_articles_sheet(ws_articles, articles)

        # ファイル保存
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = settings.output_dir / f"research_report_{timestamp}.xlsx"
        settings.ensure_directories()
        wb.save(output_path)

        logger.info(f"[{self.name}] レポート保存: {output_path}")
        return output_path

    def _create_summary_sheet(
        self,
        ws,
        analysis: Dict[str, Any],
        theme: str,
        article_count: int,
    ) -> None:
        """サマリーシートを作成"""
        # スタイル定義
        header_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")

        # タイトル
        ws["A1"] = f"リサーチレポート: {theme}"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:D1")

        ws["A2"] = f"生成日時: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A3"] = f"記事数: {article_count}件"

        # サマリー
        ws["A5"] = "概要"
        ws["A5"].font = header_font
        ws["A6"] = analysis.get("summary", "")
        ws.merge_cells("A6:D6")

        # トレンド
        ws["A8"] = "主要トレンド"
        ws["A8"].font = header_font
        for i, trend in enumerate(analysis.get("key_trends", []), start=9):
            ws[f"A{i}"] = f"• {trend}"

        # 推奨アクション
        start_row = 9 + len(analysis.get("key_trends", [])) + 1
        ws[f"A{start_row}"] = "推奨アクション"
        ws[f"A{start_row}"].font = header_font
        for i, rec in enumerate(analysis.get("recommendations", []), start=start_row + 1):
            ws[f"A{i}"] = f"• {rec}"

        # 列幅調整
        ws.column_dimensions["A"].width = 60

    def _create_articles_sheet(
        self, ws, articles: List[Dict[str, Any]]
    ) -> None:
        """記事一覧シートを作成"""
        # ヘッダー
        headers = ["No.", "タイトル", "カテゴリ", "関連度", "日本価値", "要約", "URL"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # データ
        for row, article in enumerate(articles, start=2):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=article.get("original_title", article.get("title", "")))
            ws.cell(row=row, column=3, value=article.get("category", ""))
            ws.cell(row=row, column=4, value=article.get("relevance_score", ""))
            ws.cell(row=row, column=5, value=article.get("japan_value", ""))
            ws.cell(row=row, column=6, value=article.get("summary_ja", ""))
            ws.cell(row=row, column=7, value=article.get("url", ""))

        # 列幅調整
        column_widths = [5, 50, 15, 10, 10, 40, 50]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # フィルター追加
        ws.auto_filter.ref = f"A1:G{len(articles) + 1}"
